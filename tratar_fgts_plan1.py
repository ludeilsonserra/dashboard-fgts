"""
Tratamento da aba Plan1 - Relatório FGTS (Senior/Rubi)
Autor: ChatGPT
Objetivo:
    Ler exclusivamente a aba "Plan1" do arquivo de FGTS, identificar blocos de empresa/filial,
    extrair colaboradores e linhas de totalização, e gerar bases limpas para dashboard.

Como usar no Windows:
    1) Instale as dependências:
       pip install openpyxl pandas

    2) Salve este arquivo na mesma pasta do Excel ou informe o caminho completo:
       python tratar_fgts_plan1.py "FGTS 062026.xlsx"

Saídas geradas:
    - data/base_fgts_colaboradores.csv
    - data/base_fgts_totais.csv
    - data/base_fgts_resumo_empresa.csv
    - data/base_fgts_resumo_filial.csv
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any, Optional

import pandas as pd
from openpyxl import load_workbook


# =========================
# Configurações principais
# =========================

SHEET_NAME = "Plan1"

TOTAL_KEYWORDS = (
    "Total",
    "Total Base FGTS",
    "Total a Recolher (Rubi)",
    "Total a Recolher (Sefip)",
    "Valores Dissídio",
    "Total Filial",
    "Total Empresa",
)


# =========================
# Funções auxiliares
# =========================

def parse_competencia_from_filename(file_path: str | Path) -> str:
    """
    Extrai competência do nome do arquivo.
    Exemplo:
        FGTS 062026.xlsx -> 2026-06
    """
    name = Path(file_path).stem
    match = re.search(r"(\d{2})(\d{4})", name)
    if not match:
        return ""
    mes, ano = match.groups()
    return f"{ano}-{mes}"


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def as_float(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def row_text(row: list[Any]) -> str:
    return " ".join(as_text(v) for v in row if v is not None).strip()


def detect_company(row: list[Any]) -> Optional[tuple[int, str]]:
    """
    Padrão da linha de empresa:
        col A = código numérico
        col B = "-"
        col C = nome da empresa
    """
    if len(row) >= 3 and is_number(row[0]) and as_text(row[1]) == "-" and as_text(row[2]):
        return int(row[0]), as_text(row[2])
    return None


def detect_filial(row: list[Any]) -> Optional[tuple[int, str]]:
    """
    Padrão da linha de filial:
        col A = "Filial:"
        col B = número
        col C = nome
    """
    if len(row) >= 3 and as_text(row[0]).lower() == "filial:" and is_number(row[1]):
        return int(row[1]), as_text(row[2])
    return None


def detect_total_label(row: list[Any]) -> Optional[tuple[str, int]]:
    """
    Procura o texto de totalização em qualquer coluna da linha.
    Retorna:
        (label, índice_da_coluna)
    """
    for idx, value in enumerate(row):
        text = as_text(value)
        if not text:
            continue
        if any(keyword.lower() in text.lower() for keyword in TOTAL_KEYWORDS):
            return text, idx
    return None


def first_number_after(row: list[Any], start_idx: int) -> float:
    """
    Retorna o primeiro número após a coluna do label.
    Útil para linhas como:
        Total a Recolher (Rubi): 43.85
    """
    for value in row[start_idx + 1:]:
        if is_number(value):
            return float(value)
    return 0.0


def is_employee_row(row: list[Any]) -> bool:
    """
    Linha válida de colaborador:
        col A = cadastro numérico
        col B = nome do colaborador
        col B != "-"
    """
    if len(row) < 2:
        return False

    cadastro = row[0]
    colaborador = as_text(row[1])

    if not is_number(cadastro):
        return False

    if not colaborador or colaborador == "-":
        return False

    # Proteção contra linhas de empresa ou total
    text = row_text(row).lower()
    if "total" in text or "fgts mensal" in text or "cadastro" in text:
        return False

    return True


# =========================
# Parser principal
# =========================

def parse_plan1(file_path: str | Path, sheet_name: str = SHEET_NAME) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Lê a aba Plan1 e gera:
        1) DataFrame de colaboradores
        2) DataFrame de totalizações
    """
    file_path = Path(file_path)
    competencia = parse_competencia_from_filename(file_path)

    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Aba "{sheet_name}" não encontrada. Abas disponíveis: {wb.sheetnames}')

    ws = wb[sheet_name]

    colaboradores: list[dict[str, Any]] = []
    totais: list[dict[str, Any]] = []

    current_empresa_codigo: Optional[int] = None
    current_empresa_nome: str = ""
    current_filial: Optional[int] = None
    current_filial_nome: str = ""

    for excel_row in range(1, ws.max_row + 1):
        row = [ws.cell(excel_row, col).value for col in range(1, ws.max_column + 1)]

        # 1) Detecta empresa
        company = detect_company(row)
        if company:
            empresa_codigo, empresa_nome = company

            # Quando muda de empresa, a filial deve ser redefinida.
            # Quando é apenas continuação de página da mesma empresa, preserva filial.
            if current_empresa_codigo != empresa_codigo:
                current_filial = None
                current_filial_nome = ""

            current_empresa_codigo = empresa_codigo
            current_empresa_nome = empresa_nome
            continue

        # 2) Detecta filial
        filial = detect_filial(row)
        if filial:
            current_filial, current_filial_nome = filial
            continue

        # 3) Detecta totais
        total_detected = detect_total_label(row)
        if total_detected and current_empresa_codigo is not None:
            label, label_idx = total_detected

            totais.append({
                "competencia": competencia,
                "linha_excel": excel_row,
                "empresa_codigo": current_empresa_codigo,
                "empresa_nome": current_empresa_nome,
                "filial": current_filial,
                "filial_nome": current_filial_nome,
                "tipo_total": label,
                "valor_principal": first_number_after(row, label_idx),

                # Campos estruturados para Total/Total Filial/Total Empresa
                "base_fgts_mensal": as_float(row[8]) if len(row) > 8 else 0.0,
                "valor_fgts_mensal": as_float(row[9]) if len(row) > 9 else 0.0,
                "adicional_05_mensal": as_float(row[10]) if len(row) > 10 else 0.0,
                "base_fgts_13": as_float(row[11]) if len(row) > 11 else 0.0,
                "valor_fgts_13": as_float(row[13]) if len(row) > 13 else 0.0,
                "adicional_05_13": as_float(row[14]) if len(row) > 14 else 0.0,
                "dissidio_fgts_mensal": as_float(row[15]) if len(row) > 15 else 0.0,
                "dissidio_fgts_13": as_float(row[16]) if len(row) > 16 else 0.0,
            })
            continue

        # 4) Detecta colaborador
        if is_employee_row(row) and current_empresa_codigo is not None:
            colaboradores.append({
                "competencia": competencia,
                "linha_excel": excel_row,
                "empresa_codigo": current_empresa_codigo,
                "empresa_nome": current_empresa_nome,
                "filial": current_filial,
                "filial_nome": current_filial_nome,

                "cadastro": int(row[0]),
                "colaborador": as_text(row[1]),
                "cpf": as_text(row[3]),

                # O relatório traz PIS/PASEP e CTPS em colunas com células mescladas/fragmentadas.
                # Por isso mantemos campos separados e rastreáveis.
                "pis_pasep": as_text(row[6]),
                "ctps_numero": as_text(row[7]),
                "ctps_serie": as_text(row[8]),
                "ctps_uf": as_text(row[9]),

                # Valores oficiais do FGTS mensal / 13º / dissídio
                "base_fgts_mensal": as_float(row[10]),
                "valor_fgts_mensal": as_float(row[11]),
                "adicional_05_mensal": as_float(row[12]),
                "base_fgts_13": as_float(row[13]),
                "valor_fgts_13": as_float(row[14]),
                "adicional_05_13": as_float(row[15]),
                "dissidio_fgts_mensal": as_float(row[16]),
                "dissidio_fgts_13": as_float(row[17]),

                # Indicadores derivados para dashboard
                "base_fgts_total": as_float(row[10]) + as_float(row[13]),
                "valor_fgts_total": as_float(row[11]) + as_float(row[14]) + as_float(row[16]) + as_float(row[17]),
                "tem_fgts_13": "Sim" if as_float(row[13]) > 0 or as_float(row[14]) > 0 else "Não",
                "tem_dissidio": "Sim" if as_float(row[16]) > 0 or as_float(row[17]) > 0 else "Não",
            })

    df_colaboradores = pd.DataFrame(colaboradores)
    df_totais = pd.DataFrame(totais)

    return df_colaboradores, df_totais


def gerar_resumos(df_colaboradores: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Gera resumos por empresa e filial para alimentar gráficos do dashboard.
    """
    if df_colaboradores.empty:
        return pd.DataFrame(), pd.DataFrame()

    agg_cols = {
        "cadastro": "count",
        "base_fgts_mensal": "sum",
        "valor_fgts_mensal": "sum",
        "adicional_05_mensal": "sum",
        "base_fgts_13": "sum",
        "valor_fgts_13": "sum",
        "adicional_05_13": "sum",
        "dissidio_fgts_mensal": "sum",
        "dissidio_fgts_13": "sum",
        "base_fgts_total": "sum",
        "valor_fgts_total": "sum",
    }

    resumo_empresa = (
        df_colaboradores
        .groupby(["competencia", "empresa_codigo", "empresa_nome"], dropna=False)
        .agg(agg_cols)
        .reset_index()
        .rename(columns={"cadastro": "qtd_colaboradores"})
        .sort_values("valor_fgts_total", ascending=False)
    )

    resumo_filial = (
        df_colaboradores
        .groupby(["competencia", "empresa_codigo", "empresa_nome", "filial", "filial_nome"], dropna=False)
        .agg(agg_cols)
        .reset_index()
        .rename(columns={"cadastro": "qtd_colaboradores"})
        .sort_values("valor_fgts_total", ascending=False)
    )

    return resumo_empresa, resumo_filial


def salvar_saidas(
    df_colaboradores: pd.DataFrame,
    df_totais: pd.DataFrame,
    output_dir: str | Path = "data",
) -> None:
    """
    Salva bases CSV para consumo pelo dashboard.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    resumo_empresa, resumo_filial = gerar_resumos(df_colaboradores)

    df_colaboradores.to_csv(output_dir / "base_fgts_colaboradores.csv", index=False, sep=";", encoding="utf-8-sig")
    df_totais.to_csv(output_dir / "base_fgts_totais.csv", index=False, sep=";", encoding="utf-8-sig")
    resumo_empresa.to_csv(output_dir / "base_fgts_resumo_empresa.csv", index=False, sep=";", encoding="utf-8-sig")
    resumo_filial.to_csv(output_dir / "base_fgts_resumo_filial.csv", index=False, sep=";", encoding="utf-8-sig")


def imprimir_diagnostico(df_colaboradores: pd.DataFrame, df_totais: pd.DataFrame) -> None:
    """
    Imprime resumo rápido para conferência no terminal.
    """
    print("\n=== DIAGNÓSTICO DA EXTRAÇÃO ===")
    print(f"Colaboradores extraídos: {len(df_colaboradores):,}".replace(",", "."))
    print(f"Linhas de totalização extraídas: {len(df_totais):,}".replace(",", "."))

    if not df_colaboradores.empty:
        print(f"Empresas identificadas: {df_colaboradores['empresa_codigo'].nunique()}")
        print(f"Filiais identificadas: {df_colaboradores[['empresa_codigo', 'filial']].drop_duplicates().shape[0]}")
        print(f"Base FGTS total: R$ {df_colaboradores['base_fgts_total'].sum():,.2f}")
        print(f"Valor FGTS total: R$ {df_colaboradores['valor_fgts_total'].sum():,.2f}")
        print("\nTop 5 empresas por FGTS total:")
        top = (
            df_colaboradores
            .groupby(["empresa_codigo", "empresa_nome"], dropna=False)["valor_fgts_total"]
            .sum()
            .sort_values(ascending=False)
            .head(5)
            .reset_index()
        )
        print(top.to_string(index=False))


def main() -> None:
    if len(sys.argv) < 2:
        print('Uso: python tratar_fgts_plan1.py "FGTS 062026.xlsx"')
        sys.exit(1)

    file_path = Path(sys.argv[1])

    if not file_path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    df_colaboradores, df_totais = parse_plan1(file_path)
    salvar_saidas(df_colaboradores, df_totais)
    imprimir_diagnostico(df_colaboradores, df_totais)

    print("\nArquivos gerados na pasta: data/")
    print("- base_fgts_colaboradores.csv")
    print("- base_fgts_totais.csv")
    print("- base_fgts_resumo_empresa.csv")
    print("- base_fgts_resumo_filial.csv")


if __name__ == "__main__":
    main()
